from django.http import Http404
from rest_framework import status
from rest_framework.views import APIView
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from rest_framework import authentication, permissions
from rest_framework.decorators import permission_classes
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer 
from rest_framework_simplejwt.views import TokenObtainPairView # for login page
from django.contrib.auth.hashers import check_password
from django.shortcuts import get_object_or_404
from .serializers import (
    UserSerializer, 
    UserRegisterTokenSerializer, 

)
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from rest_framework.response import Response
from .models import Video, Comment, Reply
from googleapiclient.discovery import build
import pandas as pd
import datetime
from io import BytesIO
from django.http import HttpResponse
from django.http import JsonResponse
from django.db.models import Prefetch

from django.core.serializers import serialize
import os
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import google.auth  
from googleapiclient.discovery import build  
from dotenv import load_dotenv
from .utils import extract_video_id

# register user
class UserRegisterView(APIView):
    """To Register the User"""

    def post(self, request, format=None):
        data = request.data # holds username and password (in dictionary)
        username = data["username"]
        email = data["email"]

        if username == "" or email == "":
            return Response({"detial": "username or email cannot be empty"}, status=status.HTTP_400_BAD_REQUEST)

        else:
            check_username = User.objects.filter(username=username).count()
            check_email =  User.objects.filter(email=email).count()

            if check_username:
                message = "A user with that username already exist!"
                return Response({"detail": message}, status=status.HTTP_403_FORBIDDEN)
            if check_email:
                message = "A user with that email address already exist!"
                return Response({"detail": message}, status=status.HTTP_403_FORBIDDEN)
            else:
                user = User.objects.create(
                    username=username,
                    email=email,
                    password=make_password(data["password"]),
                )
                serializer = UserRegisterTokenSerializer(user, many=False)
                return Response(serializer.data)

# login user (customizing it so that we can see fields like username, email etc as a response 
# from server, otherwise it will only provide access and refresh token)
class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    
    def validate(self, attrs):
        data = super().validate(attrs)

        serializer = UserRegisterTokenSerializer(self.user).data

        for k, v in serializer.items():
            data[k] = v
        
        return data

class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer




# get user details
class UserAccountDetailsView(APIView):

    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        try:
            user = User.objects.get(id=pk)
            serializer = UserSerializer(user, many=False)
            return Response(serializer.data, status=status.HTTP_200_OK)
            
        except:
            return Response({"details": "User not found"}, status=status.HTTP_404_NOT_FOUND)


# update user account
class UserAccountUpdateView(APIView):

    permission_classes = [permissions.IsAuthenticated]

    def put(self, request, pk):
        user = User.objects.get(id=pk)
        data = request.data

        if user:
            if request.user.id == user.id:
                user.username = data["username"]
                user.email = data["email"]

                if data["password"] != "":
                    user.password = make_password(data["password"])

                user.save()
                serializer = UserSerializer(user, many=False)
                message = {"details": "User Successfully Updated.", "user": serializer.data}
                return Response(message, status=status.HTTP_200_OK)
            else:
                return Response({"details": "Permission Denied."}, status.status.HTTP_403_FORBIDDEN)
        else:
            return Response({"details": "User not found."}, status=status.HTTP_404_NOT_FOUND)


# delete user account
class UserAccountDeleteView(APIView):

    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):

        try:
            user = User.objects.get(id=pk)
            data = request.data

            if request.user.id == user.id:
                if check_password(data["password"], user.password):
                    user.delete()
                    return Response({"details": "User successfully deleted."}, status=status.HTTP_204_NO_CONTENT)
                else:
                    return Response({"details": "Incorrect password."}, status=status.HTTP_401_UNAUTHORIZED)
            else:
                return Response({"details": "Permission Denied."}, status=status.HTTP_403_FORBIDDEN)
        except:
            return Response({"details": "User not found."}, status=status.HTTP_404_NOT_FOUND)


# get billing address (details of user address, all addresses)
class UserAddressesListView(APIView):

    def get(self, request):
        user = request.user
        user_address = BillingAddress.objects.filter(user=user)
        serializer = BillingAddressSerializer(user_address, many=True)
        
        return Response(serializer.data, status=status.HTTP_200_OK)


# get specific address only
class UserAddressDetailsView(APIView):

    def get(self, request, pk):
        user_address = BillingAddress.objects.get(id=pk)
        serializer = BillingAddressSerializer(user_address, many=False)
        return Response(serializer.data, status=status.HTTP_200_OK)


# create billing address
class CreateUserAddressView(APIView):

    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        data = request.data
        
        new_address = {
            "name": request.data["name"],
            "user": request.user.id,
            "phone_number": request.data["phone_number"],
            "pin_code": request.data["pin_code"],
            "house_no": request.data["house_no"],
            "landmark": request.data["landmark"],
            "city": request.data["city"],
            "state": request.data["state"],
        }

        serializer = BillingAddressSerializer(data=new_address, many=False)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# edit billing address
class UpdateUserAddressView(APIView):

    permission_classes = [permissions.IsAuthenticated]

    def put(self, request, pk):
        data = request.data

        try:
            user_address = BillingAddress.objects.get(id=pk)

            if request.user.id == user_address.user.id:

                updated_address = {
                    "name": data["name"] if data["name"] else user_address.name,
                    "user": request.user.id,
                    "phone_number": data["phone_number"] if data["phone_number"] else user_address.phone_number,
                    "pin_code": data["pin_code"] if data["pin_code"] else user_address.pin_code,
                    "house_no": data["house_no"] if data["house_no"] else user_address.house_no,
                    "landmark": data["landmark"] if data["landmark"] else user_address.landmark,
                    "city": data["city"] if data["city"] else user_address.city,
                    "state": data["state"] if data["state"] else user_address.state,
                }

                serializer = BillingAddressSerializer(user_address, data=updated_address)
                if serializer.is_valid():
                    serializer.save()
                    return Response(serializer.data, status=status.HTTP_200_OK)
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({"details": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)
        except:
            return Response({"details": "Not found."}, status=status.HTTP_404_NOT_FOUND)


# delete address
class DeleteUserAddressView(APIView):

    def delete(self, request, pk):
        
        try:
            user_address = BillingAddress.objects.get(id=pk)

            if request.user.id == user_address.user.id:
                user_address.delete()
                return Response({"details": "Address successfully deleted."}, status=status.HTTP_204_NO_CONTENT)
            else:
                return Response({"details": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)
        except:
            return Response({"details": "Not found."}, status=status.HTTP_404_NOT_FOUND)




# Load environment variables from .env file
load_dotenv()

# Fetch the API Key from the environment variables
API_KEY = os.getenv('API_KEY')
# for fecth the data from database

def fetch_data_videos(request):
    try:
        # Fetch videos and related comments
        videos = Video.objects.prefetch_related('comments').all()  # Adjust 'comments' to match your related name
        print('videos fetch amkit ==<<<>>>', videos)
        response_data = []
        for video in videos:
            video_data = {
                'video_id': video.video_id,
                'title': video.title,
                'description': video.description,
                'view_count': video.view_count,
                'like_count': video.like_count,
                'comment_count': video.comment_count,
                'comments': [
                    {
                        'comment_id': comment.id,  # Assuming the Comment model has 'id' as primary key
                        'text': comment.text,
                        'published_date': comment.published_date.strftime('%Y-%m-%d %H:%M:%S') if comment.published_date else None
                    }
                    for comment in video.comments.all()  # Fetch related comments
                ]
            }
            response_data.append(video_data)
        
        print('response_data ===<<>>>', response_data)
        return JsonResponse(response_data, safe=False)

    except Exception as e:
        print(f"Error in get_comments: {e}")
        return JsonResponse({'error': str(e)}, status=500)


#  insert comment and replay into database 
class FetchDataView(APIView):
    def post(self, request):
        # Get the channel URL
        channel_url = request.data.get('channel_url')
        print(f"Received channel URL: {channel_url}")

        if not channel_url:
            print("Error: Channel URL is missing.")
            return Response({'error': 'Channel URL is required.'}, status=400)

        try:
            # Extract video ID from the channel URL (assuming extract_video_id is implemented)
            video_id = extract_video_id(channel_url)
            print(f"Extracted video ID: {video_id}")

            if not video_id:
                print("Error: Invalid YouTube URL. Could not extract video ID.")
                return Response({'error': 'Invalid YouTube URL.'}, status=400)

            # Initialize YouTube API client
            youtube = build('youtube', 'v3', developerKey=API_KEY)
            print("YouTube API client initialized.")

            # Retrieve video details (title, description, etc.)
            video_response = youtube.videos().list(
                part='snippet,statistics',
                id=video_id
            ).execute()
            print(f"Video details retrieved: {video_response}")

            if not video_response['items']:
                print(f"Error: Video not found for video ID {video_id}.")
                return Response({'error': 'Video not found.'}, status=404)

            # Get video details
            video_snippet = video_response['items'][0]['snippet']
            video_statistics = video_response['items'][0].get('statistics', {})

            video_title = video_snippet['title']
            video_description = video_snippet.get('description', '')
            video_published_at = video_snippet['publishedAt']
            video_view_count = video_statistics.get('viewCount', 0)
            video_like_count = video_statistics.get('likeCount', 0)
            video_comment_count = video_statistics.get('commentCount', 0)

            # Create or get the Video object
            video, created = Video.objects.get_or_create(
                video_id=video_id,
                defaults={
                    'title': video_title,
                    'description': video_description,
                    'published_date': video_published_at,
                    'view_count': video_view_count,
                    'like_count': video_like_count,
                    'comment_count': video_comment_count,
                }
            )
            if created:
                print(f"New video record created: {video.title} (ID: {video.video_id})")
            else:
                print(f"Video already exists in the database: {video.title} (ID: {video.video_id})")

            # Retrieve video comments (latest 10 for debugging; later expand to 100 if needed)
            comment_threads_response = youtube.commentThreads().list(
                part='snippet,replies',
                videoId=video_id,
                maxResults=10  # Limit to 10 comments for debugging
            ).execute()
            print(f"Comments retrieved for video ID {video_id}: {len(comment_threads_response.get('items', []))} items.")

            # Iterate through the comments and store them
            for idx, item in enumerate(comment_threads_response.get('items', []), start=1):
                comment_snippet = item['snippet']['topLevelComment']['snippet']
                comment_id = item['id']
                comment_text = comment_snippet['textDisplay']
                comment_author = comment_snippet['authorDisplayName']
                comment_published_at = comment_snippet['publishedAt']
                comment_like_count = comment_snippet.get('likeCount', 0)

                # Create the Comment object
                comment, created = Comment.objects.get_or_create(
                video=video,
                comment_id=comment_id,
                defaults={
                    'text': comment_text,
                    'author_name': comment_author,
                    'published_date': comment_published_at,
                    'like_count': comment_like_count,
                }
                )
                
                if created:
                    print(f"Saved comment {idx}: {comment_text[:30]} (likes: {comment_like_count})")
                else:
                    print(f"Comment {idx} already exists in the database: {comment_text[:30]}")

                # If there are replies, store them
                if 'replies' in item:
                    for reply_idx, reply_item in enumerate(item['replies'].get('comments', []), start=1):
                        reply_snippet = reply_item['snippet']
                        reply_id = reply_item['id']
                        reply_text = reply_snippet['textDisplay']
                        reply_author = reply_snippet['authorDisplayName']
                        reply_published_at = reply_snippet['publishedAt']
                        reply_like_count = reply_snippet.get('likeCount', 0)

                        # Create the Reply object
                        reply, created = Reply.objects.get_or_create(
                            comment=comment,
                            reply_id=reply_id,
                            defaults={
                                'text': reply_text,
                                'author': reply_author,
                                'published_date': reply_published_at,
                                'like_count': reply_like_count,
                            }
                        )
                        if created:
                            print(f"Saved reply {reply_idx} to comment {idx}: {reply_text[:30]} (likes: {reply_like_count})")
                        else:
                            print(f"Reply {reply_idx} to comment {idx} already exists in the database.")

            print("Data fetching and saving completed successfully.")
            return Response({'message': 'Data is inserted successfully.'})

        except Exception as e:
            print(f"Error encountered: {e}")
            return Response({'error': str(e)}, status=500)



#  download excel sheet 
def export_to_excel(request):
    try:
        # Fetch video and comment data
        videos = Video.objects.all().values()
        comments = Comment.objects.all().values()

        # Create an Excel workbook
        wb = openpyxl.Workbook()

        # Helper function to adjust formatting
        def format_sheet(ws, headers, data_start_row):
            # Adjust column widths
            for col_num, header in enumerate(headers, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(len(header) + 5, 15)
            
            # Add borders to all cells with data
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Sheet 1: Video Details
        ws_videos = wb.active
        ws_videos.title = "Videos"

        video_headers = ["ID", "Video ID", "Title", "Description", "Published Date", "View Count", "Like Count", "Comment Count"]
        for col_num, header in enumerate(video_headers, 1):
            cell = ws_videos.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_num, video in enumerate(videos, 2):
            ws_videos.cell(row=row_num, column=1).value = video.get('id')
            ws_videos.cell(row=row_num, column=2).value = video.get('video_id')
            ws_videos.cell(row=row_num, column=3).value = video.get('title')
            ws_videos.cell(row=row_num, column=4).value = video.get('description')
            ws_videos.cell(row=row_num, column=5).value = video.get('published_date').strftime('%Y-%m-%d %H:%M:%S') if video.get('published_date') else None
            ws_videos.cell(row=row_num, column=6).value = video.get('view_count')
            ws_videos.cell(row=row_num, column=7).value = video.get('like_count')
            ws_videos.cell(row=row_num, column=8).value = video.get('comment_count')

        format_sheet(ws_videos, video_headers, data_start_row=2)

        # Sheet 2: Comments
        ws_comments = wb.create_sheet(title="Comments")

        comment_headers = ["ID", "Video ID", "User", "Comment Text", "Publish Date"]
        for col_num, header in enumerate(comment_headers, 1):
            cell = ws_comments.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_num, comment in enumerate(comments, 2):
            ws_comments.cell(row=row_num, column=1).value = comment.get('id')
            ws_comments.cell(row=row_num, column=2).value = comment.get('video_id')
            ws_comments.cell(row=row_num, column=4).value = comment.get('author_name')
            ws_comments.cell(row=row_num, column=4).value = comment.get('text')
            ws_comments.cell(row=row_num, column=5).value = comment.get('published_date').strftime('%Y-%m-%d %H:%M:%S') if comment.get('posted_date') else None

        format_sheet(ws_comments, comment_headers, data_start_row=2)

        # Prepare the HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="videos_and_comments.xlsx"'

        # Save the workbook to the response
        wb.save(response)
        return response

    except Exception as e:
        print("Error exporting data to Excel:", str(e))
        return HttpResponse("Internal Server Error", status=500)

